"""
Fullbay vs QuickBooks Invoice Reconciliation
Triggered by GitHub Actions via repository_dispatch.

Required env vars:
  JOB_ID                  - UUID of the fb_reconciliation_jobs row
  FULLBAY_FILE            - local path to the downloaded Fullbay CSV
  QB_FILE                 - local path to the downloaded QB Excel
  SUPABASE_URL            - Supabase project URL
  SUPABASE_SERVICE_ROLE_KEY - service role key for storage + DB updates
"""

import os, sys, json, re, tempfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

JOB_ID    = os.environ['JOB_ID']
FB_FILE   = os.environ['FULLBAY_FILE']
QB_FILE   = os.environ['QB_FILE']
SUPA_URL  = os.environ['SUPABASE_URL'].rstrip('/')
SUPA_KEY  = os.environ['SUPABASE_SERVICE_ROLE_KEY']
BUCKET    = 'fb-reconciliation-files'

HEADERS = {
    'apikey': SUPA_KEY,
    'Authorization': f'Bearer {SUPA_KEY}',
}

def update_job(patch: dict):
    r = requests.patch(
        f"{SUPA_URL}/rest/v1/fb_reconciliation_jobs?id=eq.{JOB_ID}",
        json=patch,
        headers={**HEADERS, 'Content-Type': 'application/json', 'Prefer': 'return=minimal'},
    )
    r.raise_for_status()

def upload_file(local_path: str, storage_path: str) -> str:
    with open(local_path, 'rb') as f:
        data = f.read()
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    r = requests.post(
        f"{SUPA_URL}/storage/v1/object/{BUCKET}/{storage_path}",
        data=data,
        headers={**HEADERS, 'Content-Type': content_type},
    )
    r.raise_for_status()
    return storage_path

# ── styling helpers ────────────────────────────────────────────────────────────

HDR_FILL    = PatternFill("solid", start_color="1F4E79")
HDR_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUBHDR_FILL = PatternFill("solid", start_color="2E75B6")
SUBHDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
TOTAL_FILL  = PatternFill("solid", start_color="BDD7EE")
TOTAL_FONT  = Font(bold=True, name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="D9E1F2")
MISS_FILL   = PatternFill("solid", start_color="FCE4D6")
thin        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_header(ws, row, col, text, fill=HDR_FILL, font=HDR_FONT):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill; c.font = font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c

def write_cell(ws, row, col, value, bold=False, fill=None, num_fmt=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name="Arial", size=10)
    if fill: c.fill = fill
    if num_fmt: c.number_format = num_fmt
    c.alignment = Alignment(horizontal=align)
    c.border = BORDER
    return c

def write_section_header(ws, row, label, ncols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=label)
    c.fill = SUBHDR_FILL; c.font = SUBHDR_FONT
    c.alignment = Alignment(horizontal="left")
    return row + 1

# ── main ───────────────────────────────────────────────────────────────────────

try:
    update_job({'status': 'running'})

    # ── load & clean Fullbay ──────────────────────────────────────────────────
    def strip_inv(num):
        if pd.isna(num): return None
        m = re.search(r'(\d{5})$', str(num))
        return m.group(1) if m else str(num)

    fb_raw = pd.read_csv(FB_FILE, encoding='Windows-1252')
    fb_raw['inv_num'] = fb_raw['Number'].apply(strip_inv)
    fb_raw['Sales Total'] = pd.to_numeric(fb_raw['Sales Total'], errors='coerce').fillna(0)
    fb_raw = fb_raw[fb_raw['Type'] != 'Credit Memo']
    fb_raw = fb_raw[fb_raw['inv_num'].notna()]

    fb = (fb_raw.groupby(['inv_num', 'Shop'], as_index=False)['Sales Total']
               .sum()
               .rename(columns={'Sales Total': 'FB_Total', 'Shop': 'FB_Shop'}))

    # ── load & clean QuickBooks ───────────────────────────────────────────────
    qb_raw = pd.read_excel(QB_FILE)
    qb_raw = qb_raw[qb_raw['Num'].notna()].copy()
    qb_raw['Num'] = qb_raw['Num'].astype(str).str.strip()
    qb_raw = qb_raw[~qb_raw['Num'].str.upper().str.startswith('T')]
    qb_raw = qb_raw[qb_raw['Type'] == 'Invoice']
    qb_raw['Debit']  = pd.to_numeric(qb_raw['Debit'],  errors='coerce').fillna(0)
    qb_raw['Credit'] = pd.to_numeric(qb_raw['Credit'], errors='coerce').fillna(0)
    # strip leading XX- prefix (e.g. "03-31375" → "31375")
    qb_raw['Num'] = qb_raw['Num'].str.replace(r'^\d{2}-?(?=\d{5}$)', '', regex=True)

    qb_net = (qb_raw.groupby('Num')
                    .agg(total_credit=('Credit', 'sum'), total_debit=('Debit', 'sum'))
                    .reset_index())
    qb_net['Net'] = qb_net['total_credit'] - qb_net['total_debit']

    qb_class = (qb_raw[qb_raw['Class'].notna()]
                .groupby('Num')['Class'].first().reset_index()
                .rename(columns={'Class': 'QB_Class'}))

    qb = (qb_net[['Num', 'Net']]
              .rename(columns={'Num': 'inv_num', 'Net': 'QB_Total'}))
    qb = qb.merge(qb_class.rename(columns={'Num': 'inv_num'}), on='inv_num', how='left')

    # ── merge ─────────────────────────────────────────────────────────────────
    merged = pd.merge(fb, qb, on='inv_num', how='outer')
    merged['FB_Total'] = merged['FB_Total'].fillna(0)
    merged['QB_Total'] = merged['QB_Total'].fillna(0)
    merged['Variance'] = merged['QB_Total'] - merged['FB_Total']

    in_both  = merged[(merged['FB_Total'] != 0) & (merged['QB_Total'] != 0)]
    fb_only  = merged[(merged['QB_Total'] == 0) & (merged['FB_Total'] != 0)]
    qb_only  = merged[(merged['FB_Total'] == 0) & (merged['QB_Total'] != 0)]

    fb_only_sorted = fb_only.sort_values(['FB_Shop', 'inv_num'])
    qb_only_sorted = qb_only.sort_values(['QB_Class', 'inv_num'])

    # ── build workbook ────────────────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1 – Comparison by Shop
    ws1 = wb.active
    ws1.title = "Comparison by Shop"
    ws1.freeze_panes = "A3"
    ws1.merge_cells("A1:H1")
    ws1["A1"].value = "Invoice Reconciliation – Fullbay vs QuickBooks"
    ws1["A1"].font = Font(bold=True, name="Arial", size=13, color="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Invoice #", "Fullbay Shop", "Fullbay Total ($)",
                              "QB Class", "QB Total ($)", "Variance ($)",
                              "In Fullbay", "In QuickBooks"], 1):
        set_header(ws1, 2, col, h)

    row = 3
    row = write_section_header(ws1, row, "  Invoices in Both Sources")
    alt = False
    for _, r in in_both.sort_values(['FB_Shop', 'inv_num']).iterrows():
        fill = ALT_FILL if alt else None
        write_cell(ws1, row, 1, r['inv_num'], fill=fill)
        write_cell(ws1, row, 2, r['FB_Shop'], fill=fill)
        write_cell(ws1, row, 3, r['FB_Total'], fill=fill, num_fmt='#,##0.00', align="right")
        write_cell(ws1, row, 4, r['QB_Class'], fill=fill)
        write_cell(ws1, row, 5, r['QB_Total'], fill=fill, num_fmt='#,##0.00', align="right")
        write_cell(ws1, row, 6, r['QB_Total'] - r['FB_Total'], fill=fill, num_fmt='#,##0.00', align="right")
        write_cell(ws1, row, 7, "✓", fill=fill, align="center")
        write_cell(ws1, row, 8, "✓", fill=fill, align="center")
        alt = not alt; row += 1

    write_cell(ws1, row, 2, "Subtotal – Matched", bold=True, fill=TOTAL_FILL)
    write_cell(ws1, row, 3, in_both['FB_Total'].sum(), bold=True, fill=TOTAL_FILL, num_fmt='#,##0.00', align="right")
    write_cell(ws1, row, 5, in_both['QB_Total'].sum(), bold=True, fill=TOTAL_FILL, num_fmt='#,##0.00', align="right")
    write_cell(ws1, row, 6, (in_both['QB_Total'] - in_both['FB_Total']).sum(), bold=True, fill=TOTAL_FILL, num_fmt='#,##0.00', align="right")
    for col in [1, 4, 7, 8]: write_cell(ws1, row, col, "", fill=TOTAL_FILL)
    row += 2

    row = write_section_header(ws1, row, "  In Fullbay Only (missing from QuickBooks)")
    for _, r in fb_only_sorted.iterrows():
        write_cell(ws1, row, 1, r['inv_num'], fill=MISS_FILL)
        write_cell(ws1, row, 2, r['FB_Shop'], fill=MISS_FILL)
        write_cell(ws1, row, 3, r['FB_Total'], fill=MISS_FILL, num_fmt='#,##0.00', align="right")
        write_cell(ws1, row, 4, "—", fill=MISS_FILL, align="center")
        write_cell(ws1, row, 5, "—", fill=MISS_FILL, align="center")
        write_cell(ws1, row, 6, "—", fill=MISS_FILL, align="center")
        write_cell(ws1, row, 7, "✓", fill=MISS_FILL, align="center")
        write_cell(ws1, row, 8, "✗", fill=MISS_FILL, align="center")
        row += 1
    write_cell(ws1, row, 2, "Subtotal – Fullbay Only", bold=True, fill=TOTAL_FILL)
    write_cell(ws1, row, 3, fb_only['FB_Total'].sum(), bold=True, fill=TOTAL_FILL, num_fmt='#,##0.00', align="right")
    for col in [1, 4, 5, 6, 7, 8]: write_cell(ws1, row, col, "", fill=TOTAL_FILL)
    row += 2

    row = write_section_header(ws1, row, "  In QuickBooks Only (missing from Fullbay)")
    for _, r in qb_only_sorted.iterrows():
        write_cell(ws1, row, 1, r['inv_num'], fill=MISS_FILL)
        write_cell(ws1, row, 2, "—", fill=MISS_FILL, align="center")
        write_cell(ws1, row, 3, "—", fill=MISS_FILL, align="center")
        write_cell(ws1, row, 4, r['QB_Class'], fill=MISS_FILL)
        write_cell(ws1, row, 5, r['QB_Total'], fill=MISS_FILL, num_fmt='#,##0.00', align="right")
        write_cell(ws1, row, 6, "—", fill=MISS_FILL, align="center")
        write_cell(ws1, row, 7, "✗", fill=MISS_FILL, align="center")
        write_cell(ws1, row, 8, "✓", fill=MISS_FILL, align="center")
        row += 1
    write_cell(ws1, row, 2, "Subtotal – QuickBooks Only", bold=True, fill=TOTAL_FILL)
    write_cell(ws1, row, 5, qb_only['QB_Total'].sum(), bold=True, fill=TOTAL_FILL, num_fmt='#,##0.00', align="right")
    for col in [1, 3, 4, 6, 7, 8]: write_cell(ws1, row, col, "", fill=TOTAL_FILL)
    row += 2

    grand_fill = PatternFill("solid", start_color="1F4E79")
    grand_font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
    write_cell(ws1, row, 2, "GRAND TOTAL", bold=True, fill=grand_fill)
    ws1.cell(row, 2).font = grand_font
    for col, val in [(3, merged['FB_Total'].sum()), (5, merged['QB_Total'].sum())]:
        write_cell(ws1, row, col, val, bold=True, fill=grand_fill, num_fmt='#,##0.00', align="right")
        ws1.cell(row, col).font = grand_font
    for col in [1, 4, 6, 7, 8]:
        write_cell(ws1, row, col, "", fill=grand_fill)

    for i, w in enumerate([14, 32, 20, 28, 20, 18, 14, 14], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[1].height = 24
    ws1.row_dimensions[2].height = 30

    # Sheet 2 – Discrepancies
    ws2 = wb.create_sheet("Discrepancies")
    ws2.freeze_panes = "A3"
    ws2.merge_cells("A1:E1")
    ws2["A1"].value = "Invoice Discrepancies – Missing from One Source"
    ws2["A1"].font = Font(bold=True, name="Arial", size=13, color="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Invoice #", "Shop / Class", "Total ($)", "Issue"], 1):
        c = set_header(ws2, 2, col, h)
        c.fill = PatternFill("solid", start_color="C00000")

    row2 = 3
    row2 = write_section_header(ws2, row2, "  In Fullbay but NOT in QuickBooks", ncols=4)
    for _, r in fb_only_sorted.iterrows():
        write_cell(ws2, row2, 1, r['inv_num'], fill=MISS_FILL)
        write_cell(ws2, row2, 2, r['FB_Shop'], fill=MISS_FILL)
        write_cell(ws2, row2, 3, r['FB_Total'], fill=MISS_FILL, num_fmt='#,##0.00', align="right")
        write_cell(ws2, row2, 4, "Missing from QuickBooks", fill=MISS_FILL)
        row2 += 1
    write_cell(ws2, row2, 2, "Subtotal", bold=True, fill=TOTAL_FILL)
    write_cell(ws2, row2, 3, fb_only['FB_Total'].sum(), bold=True, fill=TOTAL_FILL, num_fmt='#,##0.00', align="right")
    for col in [1, 4]: write_cell(ws2, row2, col, "", fill=TOTAL_FILL)
    row2 += 2

    row2 = write_section_header(ws2, row2, "  In QuickBooks but NOT in Fullbay", ncols=4)
    for _, r in qb_only_sorted.iterrows():
        write_cell(ws2, row2, 1, r['inv_num'], fill=MISS_FILL)
        write_cell(ws2, row2, 2, r['QB_Class'], fill=MISS_FILL)
        write_cell(ws2, row2, 3, r['QB_Total'], fill=MISS_FILL, num_fmt='#,##0.00', align="right")
        write_cell(ws2, row2, 4, "Missing from Fullbay", fill=MISS_FILL)
        row2 += 1
    write_cell(ws2, row2, 2, "Subtotal", bold=True, fill=TOTAL_FILL)
    write_cell(ws2, row2, 3, qb_only['QB_Total'].sum(), bold=True, fill=TOTAL_FILL, num_fmt='#,##0.00', align="right")
    for col in [1, 4]: write_cell(ws2, row2, col, "", fill=TOTAL_FILL)

    for i, w in enumerate([14, 32, 20, 28], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[1].height = 24

    # Sheet 3 – Summary by Shop/Class
    ws3 = wb.create_sheet("Summary by Shop-Class")
    ws3.freeze_panes = "A3"
    ws3.merge_cells("A1:F1")
    ws3["A1"].value = "Revenue Summary by Shop / Class"
    ws3["A1"].font = Font(bold=True, name="Arial", size=13, color="1F4E79")
    ws3["A1"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Shop / Class", "Source", "# Invoices", "Total Revenue ($)", "Notes"], 1):
        set_header(ws3, 2, col, h)

    row3 = 3
    fb_shop_summary = (fb.groupby('FB_Shop')
                         .agg(invoices=('inv_num', 'count'), total=('FB_Total', 'sum'))
                         .reset_index().sort_values('FB_Shop'))
    row3 = write_section_header(ws3, row3, "  Fullbay – by Shop", ncols=5)
    alt = False
    for _, r in fb_shop_summary.iterrows():
        fill = ALT_FILL if alt else None
        write_cell(ws3, row3, 1, r['FB_Shop'], fill=fill)
        write_cell(ws3, row3, 2, "Fullbay", fill=fill)
        write_cell(ws3, row3, 3, int(r['invoices']), fill=fill, align="center")
        write_cell(ws3, row3, 4, r['total'], fill=fill, num_fmt='#,##0.00', align="right")
        write_cell(ws3, row3, 5, "", fill=fill)
        alt = not alt; row3 += 1
    write_cell(ws3, row3, 1, "Fullbay Total", bold=True, fill=TOTAL_FILL)
    write_cell(ws3, row3, 3, int(fb_shop_summary['invoices'].sum()), bold=True, fill=TOTAL_FILL, align="center")
    write_cell(ws3, row3, 4, fb_shop_summary['total'].sum(), bold=True, fill=TOTAL_FILL, num_fmt='#,##0.00', align="right")
    for col in [2, 5]: write_cell(ws3, row3, col, "", fill=TOTAL_FILL)
    row3 += 2

    qb_class_summary = (qb.groupby('QB_Class')
                          .agg(invoices=('inv_num', 'count'), total=('QB_Total', 'sum'))
                          .reset_index().sort_values('QB_Class'))
    row3 = write_section_header(ws3, row3, "  QuickBooks – by Class", ncols=5)
    alt = False
    for _, r in qb_class_summary.iterrows():
        fill = ALT_FILL if alt else None
        write_cell(ws3, row3, 1, r['QB_Class'], fill=fill)
        write_cell(ws3, row3, 2, "QuickBooks", fill=fill)
        write_cell(ws3, row3, 3, int(r['invoices']), fill=fill, align="center")
        write_cell(ws3, row3, 4, r['total'], fill=fill, num_fmt='#,##0.00', align="right")
        write_cell(ws3, row3, 5, "", fill=fill)
        alt = not alt; row3 += 1
    write_cell(ws3, row3, 1, "QuickBooks Total", bold=True, fill=TOTAL_FILL)
    write_cell(ws3, row3, 3, int(qb_class_summary['invoices'].sum()), bold=True, fill=TOTAL_FILL, align="center")
    write_cell(ws3, row3, 4, qb_class_summary['total'].sum(), bold=True, fill=TOTAL_FILL, num_fmt='#,##0.00', align="right")
    for col in [2, 5]: write_cell(ws3, row3, col, "", fill=TOTAL_FILL)

    for i, w in enumerate([35, 14, 14, 20, 28], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[1].height = 24

    # ── save & upload ─────────────────────────────────────────────────────────
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        out_path = tmp.name
    wb.save(out_path)

    storage_path = f"{JOB_ID}/Invoice_Reconciliation.xlsx"
    upload_file(out_path, storage_path)
    os.unlink(out_path)

    # ── build result_json for inline display ──────────────────────────────────
    discrepancies = []
    for _, r in fb_only_sorted.iterrows():
        discrepancies.append({
            'invoice': r['inv_num'], 'status': 'fullbay_only',
            'shop': r['FB_Shop'], 'fb_total': round(float(r['FB_Total']), 2), 'qb_total': None,
        })
    for _, r in qb_only_sorted.iterrows():
        discrepancies.append({
            'invoice': r['inv_num'], 'status': 'qb_only',
            'shop': r.get('QB_Class', ''), 'fb_total': None, 'qb_total': round(float(r['QB_Total']), 2),
        })

    result_json = {
        'summary': {
            'total_fb': len(fb),
            'total_qb': len(qb),
            'in_both': len(in_both),
            'fullbay_only': len(fb_only),
            'qb_only': len(qb_only),
            'fb_total': round(float(merged['FB_Total'].sum()), 2),
            'qb_total': round(float(merged['QB_Total'].sum()), 2),
            'variance': round(float(merged['QB_Total'].sum() - merged['FB_Total'].sum()), 2),
        },
        'discrepancies': discrepancies[:200],  # cap at 200 rows for inline display
    }

    update_job({
        'status': 'done',
        'result_file_path': storage_path,
        'in_both_count': len(in_both),
        'fullbay_only_count': len(fb_only),
        'qb_only_count': len(qb_only),
        'fb_total': round(float(merged['FB_Total'].sum()), 2),
        'qb_total': round(float(merged['QB_Total'].sum()), 2),
        'variance': round(float(merged['QB_Total'].sum() - merged['FB_Total'].sum()), 2),
        'result_json': result_json,
    })

    print(f"Done. Matched: {len(in_both)}, FB only: {len(fb_only)}, QB only: {len(qb_only)}")

except Exception as e:
    update_job({'status': 'error', 'error_message': str(e)})
    print(f"ERROR: {e}", file=sys.stderr)
    sys.exit(1)
